import datetime
from openpyxl import load_workbook

#just an object wrapper for a dict
class Patient:
	def __init__(self):
		self.patient_data = {}


#returns number of (non-title) rows in a spreadsheet
def count_rows(worksheet):
	row_count = 0
	for row_cells in worksheet.iter_rows(min_row=2):
		row_count+=1
	return row_count


#returns list of empty patients for each (non-title) row in a spreadsheet
def generate_patient_list(worksheet, min_row, max_row):
	patients = []
	for row_cells in worksheet.iter_rows(min_row=min_row, max_row=max_row):
		patients.append(Patient())
	return patients


#returns list of column titles in a spreadsheet
def generate_column_title_list(worksheet):
	column_titles = []
	for row_cells in worksheet.iter_rows(min_row=1,max_row=1):
		for cell in row_cells:
			column_titles.append(cell.value)
	return column_titles


#assign worksheet data to corresponding patient in list of patients
def get_patient_data(patients, worksheet, min_row, max_row):
	for row_cells in worksheet.iter_rows(min_row=min_row, max_row=max_row):
		for cell in row_cells:
			if isinstance(cell.value, datetime.datetime):
				patients[cell.row - min_row].patient_data[f"{worksheet.cell(row=1, column=cell.column).value}"] = str(cell.value.date())
			else:
				patients[cell.row - min_row].patient_data[f"{worksheet.cell(row=1, column=cell.column).value}"] = str(cell.value)