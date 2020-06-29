HELP_TEXT = """
Functionality

This utility creates Word files (.docx) based on a given Word template for each row of a given Excel file (.xlsx). The first row is excluded and is used to identify the title of each column of data. The Word template is searched for tags that are replaced by corresponding values from the selected Excel file.


Tags

Tags are case-senstive, alpha-numeric and prepended with a "#". In order for this utility to correctly match a tag with a value from an input Excel file, the tag must match a value from the first row of the Excel file and be located inside of a Word table cell. 


Naming Convention

Title columns are identified upon selection of an Excel data file and used to populate a set of drop-down lists. Output files will be named according to the values selected in the drop-down lists and the text input into the adjacent text fields.
"""

import sys
import time

from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QFileDialog, QMessageBox

from openpyxl import load_workbook
from docx import Document

from MainWindow import Ui_MainWindow
from word import create_file
from spreadsheet import *

spreadsheet_name = ""
template_name = ""
chose_directory = ""
file_count = 0
stop_flag = False


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
	def __init__(self, *args, obj=None, **kwargs):
		super(MainWindow, self).__init__(*args, **kwargs)
		self.setupUi(self)

	#sets stop_flag upon window close, needed or else program continues to execute after window close
	def closeEvent(self, event):
		global stop_flag
		
		stop_flag = True
		super().closeEvent(event)


def stop():
	global stop_flag
	stop_flag = not stop_flag


def show_help():
	global HELP_TEXT
	
	help_window = QMessageBox()
	help_window.setText(HELP_TEXT)
	help_window.setWindowTitle("Help")
	help_window.exec_()


def maximize_range():
	window.spinBox_2.setValue(2)
	window.spinBox_3.setValue(count_rows(load_workbook(filename=spreadsheet_name).active)+1)


def update_file_count():
	global file_count
	
	window.spinBox_3.setMinimum(window.spinBox_2.value())
	window.spinBox_2.setMaximum(window.spinBox_3.value())
	file_count = window.spinBox_3.value() - window.spinBox_2.value() + 1
	window.progressBar.setFormat(f"Number of rows to be extracted: {file_count}")


def choose_input_file():
	global spreadsheet_name
	global file_count

	filename, _ = QFileDialog.getOpenFileName(filter="Excel Files(*.xlsx)")
	if filename:
		spreadsheet_name = filename
		sheet = load_workbook(filename=spreadsheet_name).active
		file_count = count_rows(sheet) 
		title_list = generate_column_title_list(sheet)
		
		#sets spinBox ranges and connects update function in order to prevent future invalid ranges
		window.spinBox_2.setMaximum(file_count+1)
		window.spinBox_2.setValue(2)
		window.spinBox_3.setMaximum(file_count+1)
		window.spinBox_3.setValue(file_count+1)
		window.spinBox_2.valueChanged.connect(update_file_count)
		window.spinBox_3.valueChanged.connect(update_file_count)
		
		if window.pushButton_6.receivers(window.pushButton_6.clicked) > 1:
			window.pushButton_6.clicked.disconnect()
		window.pushButton_6.clicked.connect(maximize_range)

		window.comboBox.clear()
		window.comboBox_2.clear()
		window.comboBox_3.clear()
		window.comboBox.addItems(title_list)
		window.comboBox_2.addItems(title_list)
		window.comboBox_3.addItems(title_list)

		#default values for comboBoxes
		if "PatFirst" in title_list and "PatLast" in title_list and "PhyFax" in title_list:
			window.comboBox.setCurrentIndex(window.comboBox_3.findText("PatFirst", QtCore.Qt.MatchFixedString))
			window.comboBox_2.setCurrentIndex(window.comboBox_3.findText("PatLast", QtCore.Qt.MatchFixedString))
			window.comboBox_3.setCurrentIndex(window.comboBox_3.findText("PhyFax", QtCore.Qt.MatchFixedString))
			window.lineEdit_2.setText("_")			

		window.label.setText(filename.split("/")[-1])
		window.progressBar.setFormat(f"Number of rows to be extracted: {file_count}")


def choose_template_filename():
	global template_name
	global file_count
	
	filename, _ = QFileDialog.getOpenFileName(filter="Word Documents(*.docx)")
	if filename:
		template_name = filename
		window.label_2.setText(filename.split("/")[-1])
		window.progressBar.setFormat(f"Number of rows to be extracted: {file_count}")


def choose_destination_directory():
	global chose_directory
	
	directory_name = QFileDialog.getExistingDirectory()
	if directory_name:
		chose_directory = directory_name + "/"
		window.label_3.setText(directory_name)


def fill_out_template():
	global file_count
	global stop_flag
	
	#makes sure that all required values have been set
	if not spreadsheet_name:
		window.progressBar.setFormat("Please select an .xlsx spreadsheet data file")
	elif not template_name:
		window.progressBar.setFormat("Please select a .docx template file")
	elif not chose_directory:
		window.progressBar.setFormat("Please select a destination directory")
	else:
		save_count = 0
		
		#locks spinBox values after run button has been clicked
		window.spinBox_2.setReadOnly(True)
		window.spinBox_3.setReadOnly(True)
		min_row = window.spinBox_2.value()
		max_row = window.spinBox_3.value()
		file_count = max_row - min_row + 1
		
		#saves default style of button before changing it to a stop button
		default_style = window.pushButton.styleSheet()
		window.pushButton.setStyleSheet("background-color:red;font-weight:bold")
		window.pushButton.setText("Stop")
		window.pushButton.clicked.disconnect()
		window.pushButton.clicked.connect(stop)

		workbook = load_workbook(filename=spreadsheet_name)
		sheet = workbook.active

		document = Document(template_name)

		patients = generate_patient_list(sheet, min_row, max_row)
		column_titles = generate_column_title_list(sheet)
		get_patient_data(patients, sheet, min_row, max_row)

		#starts performance counter and save count
		last_time = time.perf_counter()
		last_save_count = save_count

		for patient in patients:
			#allows queued events to be processed
			app.processEvents()
			
			#loop will continue even after window is closed without this
			if stop_flag:
				break
			
			new_filename = f"{chose_directory}{patient.patient_data[str(window.comboBox.currentText())]}{str(window.lineEdit.text())}{patient.patient_data[str(window.comboBox_2.currentText())]}{str(window.lineEdit_2.text())}{patient.patient_data[str(window.comboBox_3.currentText())]}{str(window.lineEdit_3.text())}.docx"
			save_count += create_file(patient.patient_data, column_titles, document, new_filename)
			current_time = time.perf_counter()
			
			#calculates file creation speed and estimated time remaining every second
			if current_time > last_time + 1:
				current_speed = save_count - last_save_count
				window.progressBar.setFormat(f"Creating {current_speed} files/s. {int((file_count-save_count)/current_speed)} seconds remaining...")
				window.progressBar.setValue((save_count/file_count)*100)
				last_save_count = save_count
				last_time = current_time
		
		#resets stop_flag if loop was broken out of
		if stop_flag:
			stop_flag = not stop_flag 
			window.progressBar.setFormat(f"Process Stopped")
			window.progressBar.setValue(0)
		else:
			window.progressBar.setFormat(f"Done")
			window.progressBar.setValue(100)

		#resets run button
		window.pushButton.setStyleSheet(default_style)
		window.pushButton.setText("Create Prescriptions")
		window.pushButton.clicked.disconnect()
		window.pushButton.clicked.connect(fill_out_template)

		#unlocks spinBoxes
		window.spinBox_2.setReadOnly(False)
		window.spinBox_3.setReadOnly(False)
			
		
app = QtWidgets.QApplication(sys.argv)

window = MainWindow()

window.setWindowTitle("Patient Data Extractor")
window.progressBar.setTextVisible(True)
window.progressBar.setFormat("Welcome!")

#needed for Windows version
window.progressBar.setAlignment(Qt.AlignCenter)

#connects buttons to appropriate functions
window.pushButton.clicked.connect(fill_out_template)
window.pushButton_2.clicked.connect(choose_input_file)
window.pushButton_3.clicked.connect(choose_template_filename)
window.pushButton_4.clicked.connect(show_help)
window.pushButton_5.clicked.connect(choose_destination_directory)


window.show()
sys.exit(app.exec_())
